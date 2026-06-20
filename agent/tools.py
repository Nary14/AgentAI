import os
import subprocess
import tempfile
import glob
import re
import json
import time
import random

# Safety blocks
BLOCKED_PATTERNS = [
    r'rm\s+-rf\s+/',
    r'rm\s+-rf\s+/\s*\*',
    r':\(\)\{\s*:\|\s*:\&\s*\};',
    r'mkfs\.',
    r'dd\s+if=.*of=/dev/sd',
    r'>\s*/dev/sda',
    r'curl.*\|\s*sh',
    r'wget.*\|\s*sh',
]

def is_safe(cmd):
    for pattern in BLOCKED_PATTERNS:
        if re.search(pattern, cmd, re.I):
            return False, pattern
    return True, ""

def run_bash(cmd, timeout=120):
    safe, reason = is_safe(cmd)
    if not safe:
        return subprocess.CompletedProcess(args=[], returncode=1, stdout="", stderr=f"BLOCKED: {reason}")
    return subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=timeout)

def exec_python(code):
    script_dir = os.path.expanduser("~/sgoinfre/AgentAI/scripts")
    os.makedirs(script_dir, exist_ok=True)
    with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, dir=script_dir) as f:
        f.write(code)
        f.flush()
        result = subprocess.run(["python3", f.name], capture_output=True, text=True, timeout=120)
        os.unlink(f.name)
    return result

def write_file(path, content):
    full = os.path.expanduser(path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    # Unescape literal \n, \t, etc. that come from AI
    try:
        content = content.encode('utf-8').decode('unicode_escape')
    except:
        pass  # If decoding fails, write as-is
    with open(full, 'w') as f:
        f.write(content)
    return f"Written: {full} ({len(content)} bytes)"

def read_file(path):
    with open(os.path.expanduser(path), 'r') as f:
        return f.read()

def list_images(folder):
    path = os.path.expanduser(folder)
    images = []
    for ext in ['*.jpg', '*.jpeg', '*.png', '*.gif', '*.webp', '*.bmp']:
        images.extend(glob.glob(os.path.join(path, ext)))
        images.extend(glob.glob(os.path.join(path, ext.upper())))
    return images

def create_excel(folder, output):
    full_folder = os.path.expanduser(folder)
    output = os.path.expanduser(output)
    
    if not os.path.exists(full_folder):
        return f"ERROR: Folder not found: {full_folder}"
    
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
    except ImportError:
        return "ERROR: openpyxl not installed. Run: pip3 install --user openpyxl"
    
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.basename(full_folder)
    
    headers = ["First Name", "Last Name", "Picture"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    
    images = list_images(full_folder)
    row = 2
    for img_path in sorted(images):
        name = os.path.splitext(os.path.basename(img_path))[0]
        parts = name.replace("_", " ").replace("-", " ").split()
        first = parts[0].title() if parts else ""
        last = " ".join(p.title() for p in parts[1:]) if len(parts) > 1 else ""
        
        ws.cell(row=row, column=1, value=first)
        ws.cell(row=row, column=2, value=last)
        
        try:
            img = XLImage(img_path)
            img.width = 120
            img.height = 120
            ws.add_image(img, f'C{row}')
            ws.row_dimensions[row].height = 100
        except Exception as e:
            ws.cell(row=row, column=3, value=f"Error: {e}")
        
        row += 1
    
    wb.save(output)
    return f"Excel created: {output} with {row-2} entries"

def download_file(url, output_path):
    try:
        import urllib.request
        output = os.path.expanduser(output_path)
        os.makedirs(os.path.dirname(output), exist_ok=True)
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as response:
            with open(output, 'wb') as f:
                f.write(response.read())
        return f"Downloaded: {output} ({os.path.getsize(output)} bytes)"
    except Exception as e:
        return f"ERROR: {str(e)}"

def web_scrape(url, action="get_text"):
    try:
        import urllib.request
        from html.parser import HTMLParser
        
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as response:
            html = response.read().decode('utf-8', errors='ignore')
        
        if action == "get_text":
            class TextExtractor(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.text = []
                    self.skip = False
                def handle_starttag(self, tag, attrs):
                    if tag in ['script', 'style', 'nav', 'footer']:
                        self.skip = True
                def handle_endtag(self, tag):
                    if tag in ['script', 'style', 'nav', 'footer']:
                        self.skip = False
                def handle_data(self, data):
                    if not self.skip:
                        self.text.append(data.strip())
            
            extractor = TextExtractor()
            extractor.feed(html)
            return ' '.join(extractor.text)[:5000]
        return html[:3000]
    except Exception as e:
        return f"ERROR: {str(e)}"

# ============== MINING TOOLS ==============

MINING_DIR = os.path.expanduser("~/sgoinfre/AgentAI/mining")
os.makedirs(MINING_DIR, exist_ok=True)
os.makedirs(f"{MINING_DIR}/logs", exist_ok=True)

MINING_PROCESSES = {}

def mine_config(path, content):
    full = os.path.expanduser(path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, 'w') as f:
        f.write(content)
    return f"Config written: {full}"

def mine_start(pool, wallet, algorithm="rx/0", threads=None):
    if "xmrig" in MINING_PROCESSES and MINING_PROCESSES["xmrig"].poll() is None:
        return "ERROR: Mining already running. Stop it first with mine_stop"
    
    temp = sys_cpu_temp()
    if isinstance(temp, (int, float)) and temp > 80:
        return f"ERROR: CPU too hot ({temp}C). Mining blocked."
    
    battery = sys_battery()
    if battery and battery.get("on_battery", False):
        return "ERROR: On battery power. Mining blocked for laptop safety."
    
    if threads is None:
        cpu_count = os.cpu_count() or 4
        threads = max(1, cpu_count // 2)
    
    xmrig_path = os.path.expanduser("~/sgoinfre/AgentAI/mining/xmrig")
    if not os.path.exists(xmrig_path):
        return f"ERROR: XMRig not found at {xmrig_path}. Run install.sh to download it."
    
    config = {
        "autosave": True,
        "cpu": {
            "enabled": True,
            "huge-pages": True,
            "hw-aes": None,
            "priority": 1,
            "memory-pool": False,
            "yield": True,
            "max-threads-hint": threads,
            "asm": True,
            "argon2-impl": None,
            "astrobwt-max-size": 550,
            "astrobwt-avx2": False,
            "cn/0": False,
            "cn-lite/0": False
        },
        "opencl": False,
        "cuda": False,
        "pools": [
            {
                "algo": algorithm,
                "coin": None,
                "url": pool,
                "user": wallet,
                "pass": "x",
                "rig-id": f"AgentAI-{os.getenv('USER')}",
                "nicehash": False,
                "keepalive": True,
                "enabled": True,
                "tls": False,
                "tls-fingerprint": None,
                "daemon": False,
                "socks5": None,
                "self-select": None,
                "submit-to-origin": False
            }
        ],
        "api": {
            "id": None,
            "worker-id": None,
            "http": {
                "enabled": True,
                "host": "127.0.0.1",
                "port": 8080,
                "access-token": None,
                "restricted": True
            }
        }
    }
    
    config_path = f"{MINING_DIR}/xmrig_active.json"
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    log_file = open(f"{MINING_DIR}/logs/mining_{int(time.time())}.log", 'w')
    proc = subprocess.Popen(
        [xmrig_path, "-c", config_path],
        stdout=log_file,
        stderr=subprocess.STDOUT,
        cwd=MINING_DIR
    )
    
    MINING_PROCESSES["xmrig"] = proc
    return f"Mining started: {algorithm} on {pool} with {threads} threads. PID: {proc.pid}"

def mine_stop():
    stopped = []
    for name, proc in list(MINING_PROCESSES.items()):
        if proc.poll() is None:
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except:
                proc.kill()
            stopped.append(name)
    
    MINING_PROCESSES.clear()
    return f"Stopped mining: {', '.join(stopped) if stopped else 'Nothing was running'}"

def mine_status():
    try:
        import requests as req
        r = req.get("http://127.0.0.1:8080/2/summary", timeout=2)
        data = r.json()
        hashrate = data.get("hashrate", {}).get("total", [0])[0]
        uptime = data.get("uptime", 0)
        threads = data.get("cpu", {}).get("threads", 0)
        return f"Hashrate: {hashrate:.2f} H/s | Threads: {threads} | Uptime: {uptime}s"
    except:
        return "Mining API not responding. Check if XMRig is running."

def mine_switch(pool, algorithm, wallet=None):
    mine_stop()
    time.sleep(1)
    wallet = wallet or "WALLET_PLACEHOLDER"
    return mine_start(pool, wallet, algorithm)

def mine_benchmark():
    xmrig_path = os.path.expanduser("~/sgoinfre/AgentAI/mining/xmrig")
    if not os.path.exists(xmrig_path):
        return "ERROR: XMRig not found"
    
    result = subprocess.run([xmrig_path, "--bench=1M"], capture_output=True, text=True, timeout=300)
    return f"Benchmark complete:\n{result.stdout[-2000:]}"

def mine_earnings(pool_api_url):
    try:
        import requests as req
        r = req.get(pool_api_url, timeout=10)
        data = r.json()
        return json.dumps(data, indent=2)[:2000]
    except Exception as e:
        return f"ERROR fetching earnings: {e}"

# ============== SYSTEM MONITORING ==============

def sys_cpu_temp():
    try:
        for path in [
            "/sys/class/thermal/thermal_zone0/temp",
            "/sys/class/hwmon/hwmon0/temp1_input",
            "/sys/class/hwmon/hwmon1/temp1_input"
        ]:
            if os.path.exists(path):
                with open(path, 'r') as f:
                    temp = int(f.read().strip()) / 1000
                    return round(temp, 1)
    except:
        pass
    return "Unknown"

def sys_load():
    try:
        with open("/proc/loadavg", 'r') as f:
            return f.read().strip()
    except:
        return "Unknown"

def sys_battery():
    try:
        for bat in glob.glob("/sys/class/power_supply/BAT*"):
            with open(f"{bat}/status", 'r') as f:
                status = f.read().strip()
            with open(f"{bat}/capacity", 'r') as f:
                capacity = int(f.read().strip())
            return {
                "on_battery": status == "Discharging",
                "status": status,
                "capacity": capacity
            }
    except:
        return None

def sys_idle_time():
    try:
        result = subprocess.run(["xprintidle"], capture_output=True, text=True, timeout=2)
        idle_ms = int(result.stdout.strip())
        return idle_ms / 1000
    except:
        return "Unknown (install xprintidle: sudo apt install xprintidle)"

def sys_is_idle(threshold_seconds=300):
    idle = sys_idle_time()
    if isinstance(idle, (int, float)):
        return idle > threshold_seconds
    return False

# ============== FAUCET & GAMBLING AUTOMATION ==============

FAUCET_LOG_DIR = os.path.expanduser("~/sgoinfre/AgentAI/faucet_logs")
os.makedirs(FAUCET_LOG_DIR, exist_ok=True)

def faucet_claim(url, claim_button_selector, cooldown=3600):
    """Automate faucet claim on a website"""
    if not browser.driver:
        return "ERROR: Browser not started. Use EXECUTE:browser_start|firefox first"
    
    try:
        browser.navigate(url)
        time.sleep(3)
        
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            
            claim_btn = WebDriverWait(browser.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, claim_button_selector))
            )
            claim_btn.click()
            time.sleep(2)
            
            page_text = browser.driver.find_element(By.TAG_NAME, "body").text
            if "success" in page_text.lower() or "claimed" in page_text.lower():
                return f"Faucet claimed successfully! Next claim in {cooldown/60} minutes"
            else:
                return f"Claim attempted. Page shows: {page_text[:200]}"
        except Exception as e:
            return f"ERROR clicking claim button: {str(e)}"
    except Exception as e:
        return f"ERROR: {str(e)}"

def auto_faucet_loop(url, claim_button, cooldown_minutes=60, max_claims=24):
    """Run faucet claims in a loop"""
    results = []
    for i in range(max_claims):
        result = faucet_claim(url, claim_button, cooldown_minutes * 60)
        results.append(f"Claim {i+1}: {result}")
        if i < max_claims - 1:
            time.sleep(cooldown_minutes * 60)
    return "\n".join(results)

def dice_roll(url, bet_amount, roll_under, api_key=None):
    """Automate dice gambling (for sites with API)"""
    return "Dice automation requires site-specific implementation and API key"

def solve_captcha(image_path):
    """Basic captcha solving placeholder"""
    return "Captcha solving requires external service API key (2captcha, Anti-Captcha)"
